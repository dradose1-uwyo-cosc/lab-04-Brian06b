# Brian Barrios Montiel
# UWYO COSC 1010
# 11/19/24
# HW5 
# Lab Section: 11
# Sources, people worked with, help given to: 
#What does it mean when I get an error code of m = ABSOLUTE_RE.match(range_string) and expected string or bytes-like object, got 'types.GenericAlias'?
#https://prohama.com/strawberry-1-pattern/ 

import openpyxl
from openpyxl.utils import get_column_letter 
from openpyxl.styles import Color, PatternFill

Colors = {"FFFF0000":['C11','C12','C13','C14','C15','D9','D10','D11','D14','D15','D16','D17','E9','E10','E11','E14','E17','E19','F11','F12','F19','F20','F21','G12','G13','G19','G20','H12','H13','H14','H17','H18','H19','H20','H23','I11','I12','I13','I14','I15','I16','I17','I18','I19','I20','I21','I22','I23','J11','J12','J13','J14','J15','J16','J18','J19','J20','J21','J22','J23','J24','K12','K15','K16','K17','K19','K22','K23','K24','L12','L15','L16','L17','L19','L22','L23','L24','M11','M12','M13','M14','M15','M16','M18','M19','M20','M21','M22','M23','M24','N11','N12','N13','N14','N15','N16','N17','N18','N19','N20','N21','N22','N23','O12','O13','O14','O17','O18','O19','O20','O23','P12','P13','P19','P20','Q11','Q12','Q13','Q19','Q20','Q21','R9','R10','R11','R14','R17','R19','S9','S10','S11','S14','S15','S16','S17','T11','T12','T13','T14','T15',],"FF00FF00":['E6','E7','F6','F7','G5','G6','G8','G9','G10','H5','H6','H7','H8','H9','H10','I5','I6','I7','I8','I9','J6','J7','K3','K4','K5','K6','K7','K8','K9','K10','L3','L4','L5','L6','L7','L8','L9','L10','M6','M7','N5','N6','N7','N8','N9','O5','O6','O7','O8','O10','P5','P6','P8','P9','P10','Q6','Q7','R6','R7'],"FF000000":['B11','B12','B13','B14','B15','C9','C10','C16','C17','D6','D7','D8','D18','D19','E5','E8','E15','E16','E20','E21','F5','F8','F9','F10','F14','F15','F16','F17','F22','G4','G7','G11','G14','G16','G17','G23','H4','H11','H15','H16','H24','I4','I10','I24','J3','J4','J5','J8','J9','J10','J25','K2','K11','K25','L2','L11','L25','M3','M4','M5','M8','M9','M10','M25','N4','N10','N24','O4','O9','O11','O15','O16','O24','P4','P7','P11','P14','P16','P17','P23','Q5','Q8','Q9','Q10','Q14','Q15','Q16','Q17','Q22','R5','R8','R15','R16','R20','R21','S6','S7','S8','S18','S19','T9','T10','T16','T17','U11','U12','U13','U14','U15','U16','U17',],"FF800080":['J17','K18','L18','M17'],"FFFFC0CB":['E18','F18','G18','P18','Q18','R18'],"FFFFFFFF":['D12','D13','E12','E13','G21','G22','H21','H22','G15','K13','K14','K20','K21','L13','L14','L20','L21','O21','O22','P15','P21','P22','R12','R13','S12','S13']}
wb = openpyxl.Workbook()
sheet = wb.active
import string
for chr in string.ascii_uppercase[:22]:
    sheet.column_dimensions[chr].width = 5
for i in range (1,26):
    sheet.row_dimensions[i].height= 13
coord = chr+str(i)





for color, cell_list in Colors.items():
    colored_filled = PatternFill(patternType="solid",fgColor=color)
    for cell in cell_list:
        sheet[cell].fill = colored_filled

wb.save("strawberry.xlsx")

