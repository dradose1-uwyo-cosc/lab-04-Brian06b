# Brian Barrios Montiel


import openpyxl
from openpyxl.utils import get_column_letter 
from openpyxl.styles import Color, PatternFill

Colors = {"FF000000":['B6','B7','B8','B9','B10','B11','B12','B13','B14','B15','C5','C16','D4','D5','D16','E4','E5','E16','F5','F9','F10','F11','F12','F16','G5','G8','G13','G16','H5','H7','H10','H11','H14','H16','I5','I7','I9','I12','I14','I16','J5','J7','J9','J12','J14','J16','K5','K7','K10','K11','K14','K16','L5','L8','L13','L16','M5','M9','M10','M11','M12','M16','N5','N16','O5','O16','P5','Q6','Q7','Q8','Q9','Q10','Q11','Q12','Q13','Q14','Q15','W7','W8','W9','W10','W11','W12','W13','W14','W15','W16','X6','X7','X8','X9','X10','X12','X13','X14','X15','X16','Y5','Y6','Y7','Y8','Y9','Y13','Y14','Y15','Y16','Z5','Z6','Z7','Z8','AA5','AA6','AA7','AA8','AA9','AA13','AA14','AA15','AA16','AB6','AB7','AB8','AB9','AB10','AB12','AB13','AB14','AB15','AB16','AC7','AC8','AC9','AC10','AC11','AC12','AC13','AC14','AC15','AC16'],"FF434343":['C8','D8','E8','F6','F7','F8','M8','N8','O8','P8','U15','U16','V6','V7','V8','W6','AD11','AD12','AD13','AD14','AD15','AD16','AE7','AE8','AE9','AE10','AE11','AE12','AE13'],"FFB7B7B7":['C6','C7','C9','C10','C11','C12','C13','C14','C15','D6','E6','G6','H6','I6','J6','K6','L6','M6','N6','O6','P6','AC5','AC6'],"FF999999":['D9','D10','D11','D12','D13','D14','D15','E9','E10','E11','E12','E13','E14','E15','F13','F14','F15','G7','G14','G15','H15','I15','J15','K15','L14','L15','M13','M14','M15','N9','N10','N11','N12','N13','N14','N15','O9','O10','O11','O12','O13','O14','O15','P9','P10','P11','P12','P13','P14','P15','AE14','AE15','AE16'],"FFF3E2F3":['D7','E7'],"FF46BDC6":['I10','I11','J10','J11'],"FF666666":['L7','M7','N7','O7','P7','U7','U8','U9','U10','U11','U12','U13','U14','V9','V10','V11','V12','V13','V14','V15','V16','W5','X4','X5','Y4','Z4','Z13','Z14','Z15','Z16','AA3','AA4','AB4','AB5','AD6','AD7','AD8','AD9','AD10'],"FFD9D9D9":['Y3','Z3','AC5','AC6'],"FFFF0000":['X11','Y10','Y12','Z9','AA10','AA12','AB11'],"FFFF9900":['Y11','Z11','Z12','AA11'],"FFFFFF00":['Z10']}
wb = openpyxl.Workbook()
sheet = wb.active
import string

for col_num in range(1,100):
    column_letter = get_column_letter(col_num)
    sheet.column_dimensions[column_letter].width = 3
for row_num in range (1,17):
    sheet.row_dimensions[row_num].height = 16
for color, cell_list in Colors.items():
    colored_filled = PatternFill(patternType="solid",fgColor=color)
    for cell in cell_list:
        sheet[cell].fill = colored_filled

wb.save("Honorsproject.xlsx")

